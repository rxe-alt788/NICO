from __future__ import annotations
import json
import urllib.request
from pathlib import Path
from typing import Any, Dict, List
from .common import csv_rows, first, json_load, read_source, safe_float

ALLOWED_PROVENANCE = {
    'ASID_AUTHORITATIVE',
    'VERIFIED_INTERIM',
    'UNVERIFIED_SUPPLEMENTARY',
}


def _pilot_mapping(location: Any, state: Any) -> tuple[str | None, str | None]:
    if str(state or '').strip().upper() != 'NSW':
        return None, None
    loc = str(location or '').lower()
    if 'palm beach' in loc:
        return 'palm-beach', 'EXACT_OR_NAMED_SITE'
    if 'north steyne' in loc or 'manly' in loc:
        return 'north-steyne', 'EXACT_OR_NAMED_SITE'
    if 'dee why' in loc:
        return 'north-steyne', 'NORTHERN_BEACHES_PILOT_PROXY'
    if 'bondi' in loc:
        return 'bondi', 'EXACT_OR_NAMED_SITE'
    if 'coogee' in loc:
        return 'coogee', 'EXACT_OR_NAMED_SITE'
    if 'cronulla' in loc:
        return 'cronulla', 'EXACT_OR_NAMED_SITE'
    if any(x in loc for x in ('vaucluse', 'shark beach', 'nielsen park', 'hermitage')):
        return 'balmoral', 'SYDNEY_HARBOUR_PILOT_PROXY'
    if 'balmoral' in loc:
        return 'balmoral', 'EXACT_OR_NAMED_SITE'
    return None, None


def _normalise(row: Dict[str, Any]) -> Dict[str, Any]:
    state = first(row, 'state')
    location = first(row, 'location', 'beach', 'site', 'incident location')
    year = first(row, 'incident.year', 'incident_year', 'year')
    month = first(row, 'incident.month', 'incident_month', 'month')
    day = first(row, 'incident.day', 'incident_day', 'day')
    time = first(row, 'incident.time', 'incident_time', 'time')
    explicit_ts = first(row, 'timestamp', 'datetime', 'incident_datetime', 'date_time')
    explicit_date = first(row, 'date', 'incident_date', 'incident date')

    timestamp = explicit_ts
    precision = 'TIMESTAMP' if explicit_ts else None
    if not timestamp and explicit_date:
        timestamp = f'{explicit_date} {time or "00:00"}'
        precision = 'DATE_TIME' if time else 'DATE'
    elif not timestamp and year and month and day:
        timestamp = f'{int(float(year)):04d}-{int(float(month)):02d}-{int(float(day)):02d} {time or "00:00"}'
        precision = 'DATE_TIME' if time else 'DATE'
    elif year and month:
        precision = 'MONTH'
    elif year:
        precision = 'YEAR'

    mapped_beach, mapping_type = _pilot_mapping(location, state)
    injury = first(row, 'victim.injury', 'victim_injury', 'injury outcome', 'fatality')
    species = first(row, 'shark.common.name', 'species', 'shark_species', 'shark species')

    return {
        'id': str(first(row, 'id', 'incident_id', 'record_id', 'case number') or ''),
        'timestamp': timestamp,
        'datePrecision': precision,
        'incidentYear': int(float(year)) if safe_float(year) is not None else None,
        'incidentMonth': int(float(month)) if safe_float(month) is not None else None,
        'state': state,
        'location': location,
        'beachId': mapped_beach,
        'pilotMappingType': mapping_type,
        'eventClass': first(row, 'event_class', 'incident_type', 'type', 'incident category') or injury or 'INCIDENT',
        'victimInjury': injury,
        'fatal': str(injury or first(row, 'fatal', 'fatality') or '').strip().lower() in ('fatal','1','true','yes','y'),
        'species': species,
        'speciesScientific': first(row, 'shark.scientific.name', 'scientific name'),
        'speciesIdentificationMethod': first(row, 'shark.identification.method', 'species identification method'),
        'speciesIdentificationSource': first(row, 'shark.identification.source', 'species identification source'),
        'speciesConfidence': first(row, 'species_confidence', 'species certainty', 'species confidence'),
        'sharkLengthM': safe_float(first(row, 'shark.length.m', 'shark length', 'length_m')),
        'latitude': safe_float(first(row, 'latitude', 'lat')),
        'longitude': safe_float(first(row, 'longitude', 'lon', 'lng')),
        'siteCategory': first(row, 'site.category', 'site category'),
        'victimActivity': first(row, 'victim.activity', 'victim activity'),
        'provokedStatus': first(row, 'provoked.unprovoked', 'provoked/unprovoked'),
        'presentAtBite': first(row, 'present.at.time.of.bite', 'present at time of bite'),
        'sourceRecord': first(row, 'source', 'source_record', 'reference'),
        'provenanceClass': 'ASID_AUTHORITATIVE',
        'validationEligible': True,
        'provenance': 'Taronga Australian Shark-Incident Database public release; pilot beach mapping is a 4NICO analytical mapping and is separately labelled'
    }


def _xlsx_rows(path: Path) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError('openpyxl is required to ingest the ASID XLSX release') from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    iterator = ws.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else '' for v in next(iterator)]
    rows = []
    for values in iterator:
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}
        if any(v not in (None, '') for v in row.values()):
            rows.append(row)
    return rows


def _download_binary(url: str, dest: Path) -> None:
    req = urllib.request.Request(url, headers={'User-Agent': '4NICO-pilot/0.7'})
    with urllib.request.urlopen(req, timeout=120) as response:
        dest.write_bytes(response.read())


def _load_interim(path: str | None) -> List[Dict[str, Any]]:
    if not path:
        return []
    p = Path(path)
    if not p.exists():
        return []
    raw = json.loads(p.read_text(encoding='utf-8'))
    records = raw.get('records', []) if isinstance(raw, dict) else []
    out = []
    for record in records:
        provenance = record.get('provenanceClass')
        if provenance not in ALLOWED_PROVENANCE:
            raise ValueError(f'Invalid provenanceClass: {provenance}')
        record = dict(record)
        record['validationEligible'] = provenance in ('ASID_AUTHORITATIVE', 'VERIFIED_INTERIM') and bool(record.get('validationEligible', True))
        out.append(record)
    return out


def load(path: str | None, url: str | None, interim_path: str | None = None) -> tuple[List[Dict[str, Any]], Dict[str, Any]]:
    source = 'not_configured'; rows: List[Dict[str, Any]] = []
    if path:
        p = Path(path); source = f'file:{p}'
        if not p.exists():
            return [], {'state': 'MISSING', 'source': source, 'note': 'Configured ASID file does not exist.'}
        if p.suffix.lower() in ('.xlsx', '.xlsm'):
            rows = _xlsx_rows(p)
        else:
            text = p.read_text(encoding='utf-8-sig'); stripped = text.lstrip()
            raw = json_load(text) if stripped.startswith(('{','[')) else csv_rows(text)
            rows = raw.get('incidents', []) if isinstance(raw, dict) else raw
    elif url:
        source = url
        if '.xlsx' in url.lower() or '.xlsm' in url.lower():
            temp = Path('/tmp/4nico_asid_public.xlsx'); _download_binary(url, temp); rows = _xlsx_rows(temp)
        else:
            text, _ = read_source(None, url)
            if text:
                stripped = text.lstrip(); raw = json_load(text) if stripped.startswith(('{','[')) else csv_rows(text)
                rows = raw.get('incidents', []) if isinstance(raw, dict) else raw

    incidents = [_normalise(r) for r in rows if isinstance(r, dict)]
    incidents = [i for i in incidents if i['timestamp'] or i['location']]
    interim = _load_interim(interim_path)

    # Authoritative records win on exact ID collision. Interim/supplementary rows fill
    # post-release gaps without being relabelled as ASID data.
    authoritative_ids = {i.get('id') for i in incidents if i.get('id')}
    incidents.extend(i for i in interim if not i.get('id') or i.get('id') not in authoritative_ids)

    nsw_2026 = sum(1 for i in incidents if str(i.get('state') or '').upper() == 'NSW' and (i.get('incidentYear') == 2026 or str(i.get('timestamp') or '').startswith('2026-')))
    eligible = sum(1 for i in incidents if i.get('validationEligible'))
    counts = {p: sum(1 for i in incidents if i.get('provenanceClass') == p) for p in sorted(ALLOWED_PROVENANCE)}
    state = 'LOADED' if rows else ('INTERIM_ONLY' if interim else 'MISSING')
    return incidents, {
        'state': state,
        'source': source,
        'count': len(incidents),
        'validationEligibleCount': eligible,
        'provenanceCounts': counts,
        'nsw2026Count': nsw_2026,
        'sourceClass': 'MIXED_INCIDENT_LEDGER' if interim else 'ASID_AUTHORITATIVE',
        'note': 'Validation metrics accept ASID_AUTHORITATIVE and VERIFIED_INTERIM only. UNVERIFIED_SUPPLEMENTARY records remain visible but excluded.'
    }
